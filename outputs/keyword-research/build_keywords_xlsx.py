from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path

headers = [
    "Keyword", "Idioma", "Intento", "Funnel", "Segmento", "Match type sugerido",
    "Ad group", "Landing sugerida", "Ángulo de copy", "Prioridad", "Notas"
]

product_rows = [
("TerminalSync", "Brand", "Navigational", "BOFU", "Dev/Founder", "Exact", "Brand", "/", "Sincroniza tus agentes IA en cualquier Mac", 1, "Defender marca propia."),
("terminalsync.ai", "Brand", "Navigational", "BOFU", "Dev/Founder", "Exact", "Brand", "/", "Página oficial de TerminalSync", 1, "Variación dominio."),
("Terminal Sync", "Brand", "Navigational", "BOFU", "Dev/Founder", "Exact", "Brand", "/", "Tu terminal IA, sincronizada", 1, "Variación con espacio."),
("TerminalSyn AI", "Brand", "Navigational", "BOFU", "Dev/Founder", "Phrase", "Brand typos", "/", "Encuentra TerminalSync.ai", 2, "Typo probable del usuario."),
("TerminalSync app", "Brand", "Navigational", "BOFU", "Dev/Founder", "Phrase", "Brand", "/", "Descarga la app", 1, "Alta intención."),
("TerminalSync download", "EN", "Transactional", "BOFU", "Developer", "Phrase", "Download", "/download", "Start syncing AI agents", 1, "Keyword de descarga."),
("descargar TerminalSync", "ES", "Transactional", "BOFU", "Developer", "Phrase", "Download ES", "/es/download", "Descargá TerminalSync", 1, "LatAm/España."),
("sync Claude Code and Codex", "EN", "Problem-aware", "MOFU", "AI coder", "Phrase", "Agent Sync", "/", "Claude, Codex and Gemini share context", 1, "Core pain."),
("Claude Code Codex sync", "EN", "Problem-aware", "MOFU", "AI coder", "Phrase", "Agent Sync", "/", "Stop repeating context", 1, "Alta relevancia."),
("Claude Code Gemini Codex sync", "EN", "Problem-aware", "MOFU", "AI coder", "Phrase", "Agent Sync", "/", "One memory across agents", 1, "Trío exacto del producto."),
("sync AI coding agents", "EN", "Commercial", "MOFU", "Developer", "Phrase", "Agent Sync", "/", "Keep every coding agent in sync", 1, "Keyword genérico fuerte."),
("AI agent memory sync", "EN", "Commercial", "MOFU", "Developer", "Phrase", "Memory Sync", "/", "Shared memory for local AI agents", 1, "Core positioning."),
("shared memory for AI agents", "EN", "Commercial", "MOFU", "Developer", "Phrase", "Memory Sync", "/", "One shared memory layer", 1, "Problema semántico."),
("Claude Code shared memory", "EN", "Commercial", "MOFU", "Claude users", "Phrase", "Claude Code", "/", "Claude remembers across machines", 1, "Muy relevante."),
("Codex shared memory", "EN", "Commercial", "MOFU", "Codex users", "Phrase", "Codex", "/", "Codex keeps project context", 1, "Muy relevante."),
("Gemini CLI shared memory", "EN", "Commercial", "MOFU", "Gemini users", "Phrase", "Gemini", "/", "Gemini CLI with synced context", 2, "Usuario de Gemini CLI."),
("Claude Code across machines", "EN", "Problem-aware", "MOFU", "Claude users", "Phrase", "Claude Code", "/", "Same Claude context on laptop and desktop", 1, "Dolor multi-máquina."),
("Codex across machines", "EN", "Problem-aware", "MOFU", "Codex users", "Phrase", "Codex", "/", "Your Codex setup follows you", 1, "Dolor multi-máquina."),
("AI coding agent across machines", "EN", "Problem-aware", "MOFU", "Developer", "Phrase", "Multi-machine", "/", "Move from laptop to tower without reset", 1, "Core use case."),
("sync AI agents across computers", "EN", "Problem-aware", "MOFU", "Developer", "Phrase", "Multi-machine", "/", "Every computer gets the same agent setup", 1, "Core use case."),
("keep AI agents in sync", "EN", "Problem-aware", "MOFU", "Developer", "Phrase", "Agent Sync", "/", "Never reconfigure agents again", 1, "Good ad headline."),
("AI terminal sync", "EN", "Commercial", "MOFU", "Developer", "Phrase", "Terminal Sync", "/", "Your terminal state, encrypted and synced", 1, "Category-like."),
("terminal sync for AI agents", "EN", "Commercial", "MOFU", "Developer", "Phrase", "Terminal Sync", "/", "Terminal workflows that follow you", 1, "Category-like."),
("terminal session sync", "EN", "Commercial", "MOFU", "Developer", "Phrase", "Terminal Sync", "/", "Sync sessions, tools and memory", 2, "May include generic terminal users."),
("terminal workspace sync", "EN", "Commercial", "MOFU", "Developer", "Phrase", "Terminal Sync", "/", "A portable AI terminal workspace", 2, "Product category."),
("AI terminal workspace", "EN", "Commercial", "TOFU", "Developer", "Broad", "AI Terminal", "/", "Workspace for Claude, Codex and Gemini", 2, "Broad but relevant."),
("AI terminal for developers", "EN", "Commercial", "TOFU", "Developer", "Broad", "AI Terminal", "/", "Built for agentic coding workflows", 2, "Competitive category."),
("terminal AI assistant", "EN", "Commercial", "TOFU", "Developer", "Broad", "AI Terminal", "/", "Make your terminal AI-ready", 3, "Más amplio/competitivo."),
("AI coding terminal", "EN", "Commercial", "TOFU", "Developer", "Broad", "AI Terminal", "/", "Run AI coding agents with persistent context", 2, "Relevant broad."),
("agentic coding terminal", "EN", "Commercial", "TOFU", "AI developer", "Phrase", "AI Terminal", "/", "For serious agentic coding workflows", 2, "Emerging category."),
("multi agent terminal", "EN", "Commercial", "MOFU", "Power user", "Phrase", "Multi-agent", "/", "Coordinate multiple agents from terminal", 2, "Adjacent competitors."),
("multi AI terminal", "EN", "Commercial", "MOFU", "Power user", "Phrase", "Multi-agent", "/", "Claude, Codex, Gemini in sync", 2, "Adjacent category."),
("Claude Code workflow", "EN", "Informational", "TOFU", "Claude users", "Phrase", "Claude Code", "/blog", "Upgrade your Claude Code workflow", 2, "Use for content/remarketing."),
("Claude Code setup", "EN", "Informational", "TOFU", "Claude users", "Phrase", "Claude Code", "/blog", "Setup once, sync everywhere", 2, "High search likely, broad."),
("Claude Code config sync", "EN", "Problem-aware", "MOFU", "Claude users", "Phrase", "Config Sync", "/", "Sync your Claude config safely", 1, "Very product-relevant."),
("Codex config sync", "EN", "Problem-aware", "MOFU", "Codex users", "Phrase", "Config Sync", "/", "Sync Codex settings across Macs", 1, "Very product-relevant."),
("Gemini CLI config sync", "EN", "Problem-aware", "MOFU", "Gemini users", "Phrase", "Config Sync", "/", "Gemini CLI setup that follows you", 2, "Relevant."),
("MCP config sync", "EN", "Problem-aware", "MOFU", "MCP users", "Phrase", "MCP Sync", "/", "Sync MCP servers across agents", 1, "Marketplace bridge."),
("sync MCP servers", "EN", "Commercial", "MOFU", "MCP users", "Phrase", "MCP Sync", "/marketplace", "Install once, use everywhere", 1, "Marketplace/product overlap."),
("MCP servers across machines", "EN", "Problem-aware", "MOFU", "MCP users", "Phrase", "MCP Sync", "/marketplace", "Your MCP stack follows you", 1, "Strong pain."),
("Claude MCP sync", "EN", "Problem-aware", "MOFU", "Claude users", "Phrase", "MCP Sync", "/marketplace", "Keep Claude MCP tools synced", 1, "High fit."),
("Codex MCP sync", "EN", "Problem-aware", "MOFU", "Codex users", "Phrase", "MCP Sync", "/marketplace", "Keep Codex tools synced", 1, "High fit."),
("encrypted AI agent memory", "EN", "Security", "MOFU", "Security-conscious dev", "Phrase", "Security", "/security", "Encrypted by default", 1, "Differentiator."),
("encrypted terminal sync", "EN", "Security", "MOFU", "Security-conscious dev", "Phrase", "Security", "/security", "Sync without leaking secrets", 1, "Core security."),
("sync env files encrypted", "EN", "Problem-aware", "MOFU", "Developer", "Phrase", "Env Vault", "/", "Encrypted env vault for projects", 2, "Feature keyword."),
("developer env vault", "EN", "Commercial", "MOFU", "Developer", "Phrase", "Env Vault", "/", "Protect local project secrets", 2, "Feature keyword."),
("sync .env across machines", "EN", "Problem-aware", "MOFU", "Developer", "Phrase", "Env Vault", "/", "Encrypted .env sync for dev machines", 2, "Strong pain, security caveat."),
("sync terminal secrets", "EN", "Problem-aware", "MOFU", "Developer", "Phrase", "Security", "/security", "Secrets sync without plaintext Drive", 2, "Security angle."),
("AI agent context management", "EN", "Commercial", "MOFU", "AI developer", "Phrase", "Context", "/", "Context that persists across tools", 1, "Core concept."),
("persistent context for AI agents", "EN", "Commercial", "MOFU", "AI developer", "Phrase", "Context", "/", "Persistent memory for coding agents", 1, "Core concept."),
("persistent Claude Code context", "EN", "Problem-aware", "MOFU", "Claude users", "Phrase", "Context", "/", "Claude keeps context between sessions", 1, "Strong fit."),
("persistent Codex context", "EN", "Problem-aware", "MOFU", "Codex users", "Phrase", "Context", "/", "Codex remembers your project", 1, "Strong fit."),
("AI coding agents forget context", "EN", "Problem-aware", "TOFU", "AI coder", "Phrase", "Pain", "/", "Stop repeating yourself to every agent", 1, "Pain-led search."),
("stop repeating context to AI", "EN", "Problem-aware", "TOFU", "AI coder", "Phrase", "Pain", "/", "Tell it once, every agent knows", 1, "Pain-led ad."),
("Claude Code forgets context", "EN", "Problem-aware", "TOFU", "Claude users", "Phrase", "Pain", "/", "Give Claude persistent project memory", 2, "Likely support/search pain."),
("Codex forgets context", "EN", "Problem-aware", "TOFU", "Codex users", "Phrase", "Pain", "/", "Give Codex persistent project memory", 2, "Likely support/search pain."),
("AI coding agent setup", "EN", "Informational", "TOFU", "Developer", "Phrase", "Setup", "/blog", "A better setup for AI coding agents", 2, "Content/ad combo."),
("best AI coding agent setup", "EN", "Commercial", "TOFU", "Developer", "Phrase", "Setup", "/blog", "Claude + Codex + Gemini, synced", 2, "Comparative content."),
("AI developer workflow tools", "EN", "Commercial", "TOFU", "Developer", "Broad", "Workflow", "/", "Workflow tools for agentic devs", 3, "Broad."),
("agentic developer tools", "EN", "Commercial", "TOFU", "Developer", "Broad", "Workflow", "/", "Tools for agentic coding", 3, "Broad."),
("vibe coding tools", "EN", "Commercial", "TOFU", "Founder/Dev", "Broad", "Vibe Coding", "/", "Make vibe coding less chaotic", 2, "Popular phrase, broad."),
("vibe coding workflow", "EN", "Commercial", "TOFU", "Founder/Dev", "Phrase", "Vibe Coding", "/", "Keep agents aligned while vibe coding", 2, "Good awareness."),
("Claude Code alternative", "EN", "Commercial", "MOFU", "Claude users", "Phrase", "Competitor", "/vs/claude-code", "Not an alternative—your sync layer", 3, "Must be precise to avoid mismatch."),
("Codex alternative", "EN", "Commercial", "MOFU", "Codex users", "Phrase", "Competitor", "/vs/codex", "Use Codex with better continuity", 3, "Could be expensive."),
("Warp AI alternative", "EN", "Commercial", "MOFU", "Terminal users", "Phrase", "Competitor", "/vs/warp", "TerminalSync is the sync layer for agents", 3, "Competitor page if exists."),
("tmux for AI agents", "EN", "Commercial", "MOFU", "Power user", "Phrase", "Adjacent", "/", "Beyond panes: memory and config sync", 2, "Adjacent category."),
("remote AI coding agents", "EN", "Commercial", "TOFU", "Developer", "Broad", "Remote", "/", "Run locally, continue anywhere", 3, "Adjacent."),
("AI agents on multiple Macs", "EN", "Problem-aware", "MOFU", "Mac developers", "Phrase", "Mac", "/", "Same AI setup on every Mac", 1, "Mac-specific."),
("Mac AI coding workflow", "EN", "Commercial", "TOFU", "Mac developers", "Phrase", "Mac", "/", "Agentic coding workflow for Mac", 2, "Good fit if macOS app."),
("sincronizar Claude Code y Codex", "ES", "Problem-aware", "MOFU", "Dev ES", "Phrase", "Agent Sync ES", "/es", "Claude, Codex y Gemini con la misma memoria", 1, "Keyword core en español."),
("sincronizar agentes IA", "ES", "Commercial", "MOFU", "Dev ES", "Phrase", "Agent Sync ES", "/es", "Agentes IA sincronizados entre máquinas", 1, "Core ES."),
("memoria compartida agentes IA", "ES", "Commercial", "MOFU", "Dev ES", "Phrase", "Memory ES", "/es", "Una memoria para Claude, Codex y Gemini", 1, "Core ES."),
("memoria persistente Claude Code", "ES", "Problem-aware", "MOFU", "Claude ES", "Phrase", "Claude ES", "/es", "Claude recuerda tu proyecto", 1, "Core ES."),
("memoria persistente Codex", "ES", "Problem-aware", "MOFU", "Codex ES", "Phrase", "Codex ES", "/es", "Codex recuerda tu proyecto", 1, "Core ES."),
("sincronizar MCP servers", "ES", "Problem-aware", "MOFU", "MCP ES", "Phrase", "MCP ES", "/es/marketplace", "Instalá MCP una vez y usalo en todos lados", 1, "Marketplace/product overlap."),
("terminal IA para programadores", "ES", "Commercial", "TOFU", "Dev ES", "Broad", "AI Terminal ES", "/es", "Tu terminal para trabajar con agentes IA", 2, "Awareness ES."),
("herramientas vibe coding", "ES", "Commercial", "TOFU", "Founder/Dev ES", "Broad", "Vibe ES", "/es", "Vibe coding sin perder contexto", 2, "Awareness ES."),
("agentes IA para programar", "ES", "Informational", "TOFU", "Dev ES", "Broad", "AI Coding ES", "/es/blog", "Cómo trabajar con Claude, Codex y Gemini", 3, "Broad Spanish."),
("configuración Claude Code", "ES", "Informational", "TOFU", "Claude ES", "Phrase", "Claude ES", "/es/blog", "Configurá Claude Code y sincronizalo", 2, "Content."),
("Claude Code varias computadoras", "ES", "Problem-aware", "MOFU", "Claude ES", "Phrase", "Multi-machine ES", "/es", "Claude listo en laptop y desktop", 1, "Strong ES pain."),
("Codex varias computadoras", "ES", "Problem-aware", "MOFU", "Codex ES", "Phrase", "Multi-machine ES", "/es", "Codex listo en laptop y desktop", 1, "Strong ES pain."),
("sincronizar .env cifrado", "ES", "Security", "MOFU", "Dev ES", "Phrase", "Security ES", "/es/security", "Secrets cifrados entre máquinas", 2, "Feature ES."),
]

market_rows = [
("TerminalSync marketplace", "Brand", "Navigational", "BOFU", "Marketplace buyer", "Exact", "Brand Marketplace", "/marketplace", "Browse TerminalSync connectors and skills", 1, "Defender marketplace own brand."),
("TerminalSync connectors", "Brand", "Navigational", "BOFU", "Marketplace buyer", "Phrase", "Brand Marketplace", "/marketplace", "Install agent connectors", 1, "Own category."),
("TerminalSync skills", "Brand", "Navigational", "BOFU", "Marketplace buyer", "Phrase", "Brand Marketplace", "/marketplace/skills", "Skills for Claude, Codex and Gemini", 1, "Own category."),
("MCP marketplace", "EN", "Commercial", "MOFU", "MCP users", "Phrase", "MCP Marketplace", "/marketplace", "Find MCP servers for your AI agents", 1, "Core marketplace keyword."),
("MCP server marketplace", "EN", "Commercial", "MOFU", "MCP users", "Phrase", "MCP Marketplace", "/marketplace", "Browse, install and sync MCP servers", 1, "Core marketplace keyword."),
("AI agent marketplace", "EN", "Commercial", "MOFU", "AI builders", "Phrase", "AI Marketplace", "/marketplace", "Connectors, skills and CLI tools for agents", 1, "Core marketplace keyword."),
("AI tools marketplace for developers", "EN", "Commercial", "TOFU", "Developers", "Phrase", "AI Marketplace", "/marketplace", "Developer-first AI tool marketplace", 2, "Broad but relevant."),
("Claude Code marketplace", "EN", "Commercial", "MOFU", "Claude users", "Phrase", "Claude Marketplace", "/marketplace", "Tools and skills for Claude Code", 1, "High relevance."),
("Codex marketplace", "EN", "Commercial", "MOFU", "Codex users", "Phrase", "Codex Marketplace", "/marketplace", "Tools and skills for Codex", 1, "High relevance."),
("Gemini CLI marketplace", "EN", "Commercial", "MOFU", "Gemini users", "Phrase", "Gemini Marketplace", "/marketplace", "Tools and skills for Gemini CLI", 2, "Relevant."),
("AI agent connectors", "EN", "Commercial", "MOFU", "AI users", "Phrase", "Connectors", "/marketplace/connectors", "Connect your agent to the tools you use", 1, "Core connector category."),
("MCP connectors", "EN", "Commercial", "MOFU", "MCP users", "Phrase", "Connectors", "/marketplace/connectors", "MCP connectors synced across agents", 1, "Core."),
("Claude connectors", "EN", "Commercial", "MOFU", "Claude users", "Phrase", "Connectors", "/marketplace/connectors", "Give Claude access to your apps", 1, "High fit."),
("Codex connectors", "EN", "Commercial", "MOFU", "Codex users", "Phrase", "Connectors", "/marketplace/connectors", "Give Codex access to your apps", 1, "High fit."),
("AI agent skills", "EN", "Commercial", "MOFU", "AI users", "Phrase", "Skills", "/marketplace/skills", "Reusable skills for agent workflows", 1, "Core skills category."),
("Claude Code skills", "EN", "Commercial", "MOFU", "Claude users", "Phrase", "Skills", "/marketplace/skills", "Install reusable Claude Code skills", 1, "High relevance."),
("Codex skills", "EN", "Commercial", "MOFU", "Codex users", "Phrase", "Skills", "/marketplace/skills", "Install reusable Codex skills", 1, "High relevance."),
("Gemini CLI skills", "EN", "Commercial", "MOFU", "Gemini users", "Phrase", "Skills", "/marketplace/skills", "Skills for Gemini CLI workflows", 2, "Relevant."),
("Claude Code plugins", "EN", "Commercial", "MOFU", "Claude users", "Phrase", "Plugins", "/marketplace", "Plugins that follow your agent setup", 1, "Users may say plugin not MCP."),
("Codex plugins", "EN", "Commercial", "MOFU", "Codex users", "Phrase", "Plugins", "/marketplace", "Plugins that follow your Codex setup", 1, "Users may say plugin."),
("AI coding plugins", "EN", "Commercial", "TOFU", "Developers", "Broad", "Plugins", "/marketplace", "Plugins for agentic coding", 2, "Broad."),
("MCP servers for Claude", "EN", "Commercial", "MOFU", "Claude users", "Phrase", "MCP by Agent", "/marketplace/connectors", "Best MCP servers for Claude, synced", 1, "Strong SEO/ad query."),
("MCP servers for Codex", "EN", "Commercial", "MOFU", "Codex users", "Phrase", "MCP by Agent", "/marketplace/connectors", "MCP servers for Codex, synced", 1, "Strong."),
("MCP servers for Gemini", "EN", "Commercial", "MOFU", "Gemini users", "Phrase", "MCP by Agent", "/marketplace/connectors", "MCP servers for Gemini CLI", 2, "Strong."),
("best MCP servers", "EN", "Commercial", "TOFU", "MCP users", "Phrase", "MCP Discovery", "/marketplace/connectors", "Curated MCP servers for real workflows", 2, "SEO first, paid test cautiously."),
("MCP server list", "EN", "Informational", "TOFU", "MCP users", "Phrase", "MCP Discovery", "/marketplace/connectors", "A practical list of MCP servers", 2, "Content + retargeting."),
("MCP directory", "EN", "Commercial", "TOFU", "MCP users", "Phrase", "MCP Discovery", "/marketplace/connectors", "Search MCP servers by workflow", 2, "Directory term."),
("AI agent directory", "EN", "Commercial", "TOFU", "AI users", "Phrase", "AI Discovery", "/marketplace", "Discover agent-ready tools", 3, "Broad."),
("developer tools marketplace", "EN", "Commercial", "TOFU", "Developers", "Broad", "AI Marketplace", "/marketplace", "A marketplace for AI-native developer tools", 3, "Very broad."),
("publish MCP server", "EN", "Transactional", "BOFU", "MCP creators", "Phrase", "Publishers", "/marketplace/submit", "Publish your MCP server", 1, "Supply-side acquisition."),
("sell MCP server", "EN", "Transactional", "BOFU", "MCP creators", "Phrase", "Publishers", "/marketplace/submit", "Sell tools to AI developers", 1, "Publisher revenue intent."),
("sell AI developer tools", "EN", "Transactional", "BOFU", "Tool creators", "Phrase", "Publishers", "/marketplace/submit", "Reach developers using AI agents", 1, "Supply-side."),
("publish AI agent skill", "EN", "Transactional", "BOFU", "Skill creators", "Phrase", "Publishers", "/marketplace/submit", "Publish skills for Claude/Codex/Gemini", 1, "Supply-side."),
("AI skill marketplace", "EN", "Commercial", "MOFU", "Skill creators/buyers", "Phrase", "Skills", "/marketplace/skills", "Buy and publish AI agent skills", 1, "Two-sided."),
("Claude skill marketplace", "EN", "Commercial", "MOFU", "Claude users", "Phrase", "Skills", "/marketplace/skills", "Claude skills you can install", 1, "High fit."),
("Codex skill marketplace", "EN", "Commercial", "MOFU", "Codex users", "Phrase", "Skills", "/marketplace/skills", "Codex skills you can install", 1, "High fit."),
("AI prompt marketplace for developers", "EN", "Commercial", "TOFU", "Developers", "Phrase", "Skills", "/marketplace/skills", "Reusable prompts packaged as skills", 2, "Adjacent to skills."),
("prompt packs for Claude Code", "EN", "Commercial", "MOFU", "Claude users", "Phrase", "Skills", "/marketplace/skills", "Claude Code prompt packs and skills", 2, "Adjacent."),
("code review AI skill", "EN", "Transactional", "MOFU", "Developers", "Phrase", "Skill Use Cases", "/marketplace/skills/code-reviewer", "Install a code review skill", 1, "Existing skill."),
("SEO AI skill", "EN", "Transactional", "MOFU", "Marketers/devs", "Phrase", "Skill Use Cases", "/marketplace/skills/seo-auditor", "Install an SEO auditor skill", 2, "Existing skill."),
("copywriting AI skill", "EN", "Transactional", "MOFU", "Marketers", "Phrase", "Skill Use Cases", "/marketplace/skills/copywriter", "Install a direct-response copy skill", 2, "Existing skill."),
("email drafter AI skill", "EN", "Transactional", "MOFU", "Operators", "Phrase", "Skill Use Cases", "/marketplace/skills/email-drafter", "Install an email drafting skill", 2, "Existing skill."),
("Meta ads AI skill", "EN", "Transactional", "MOFU", "Marketers", "Phrase", "Skill Use Cases", "/marketplace/skills/meta-ads-creator", "Install a Meta ads creator skill", 2, "Existing skill."),
("Google Drive MCP", "EN", "Transactional", "MOFU", "Knowledge workers", "Phrase", "Connector Apps", "/marketplace/connectors/google-drive", "Let your agent read Drive", 1, "Existing connector."),
("Gmail MCP", "EN", "Transactional", "MOFU", "Operators", "Phrase", "Connector Apps", "/marketplace/connectors/gmail", "Your inbox, talking to AI", 1, "Existing connector."),
("Notion MCP", "EN", "Transactional", "MOFU", "Teams/founders", "Phrase", "Connector Apps", "/marketplace/connectors/notion", "Make your AI understand Notion", 1, "Existing connector."),
("Airtable MCP", "EN", "Transactional", "MOFU", "Operators", "Phrase", "Connector Apps", "/marketplace/connectors/airtable", "Airtable inside your agent", 2, "Existing connector."),
("Supabase MCP", "EN", "Transactional", "MOFU", "Developers", "Phrase", "Connector Apps", "/marketplace/connectors/supabase", "Query Supabase from your agent", 1, "Existing connector."),
("Vercel MCP", "EN", "Transactional", "MOFU", "Developers", "Phrase", "Connector Apps", "/marketplace/connectors/vercel", "Deploy workflows from your agent", 1, "Existing connector."),
("Webflow MCP", "EN", "Transactional", "MOFU", "Agencies", "Phrase", "Connector Apps", "/marketplace/connectors/webflow", "Let AI edit Webflow", 2, "Existing connector."),
("Framer MCP", "EN", "Transactional", "MOFU", "Designers/founders", "Phrase", "Connector Apps", "/marketplace/connectors/framer", "Let AI edit Framer", 2, "Existing connector."),
("WhatsApp MCP", "EN", "Transactional", "MOFU", "Businesses", "Phrase", "Connector Apps", "/marketplace/connectors/whatsapp", "WhatsApp Business with AI replies", 2, "Existing connector."),
("Make MCP", "EN", "Transactional", "MOFU", "Automation users", "Phrase", "Connector Apps", "/marketplace/connectors/make", "Let AI trigger automations", 2, "Existing connector."),
("Pipedream MCP", "EN", "Transactional", "MOFU", "Automation builders", "Phrase", "Connector Apps", "/marketplace/connectors/pipedream", "Connect AI to thousands of APIs", 2, "Existing connector."),
("Typeform MCP", "EN", "Transactional", "MOFU", "Researchers/ops", "Phrase", "Connector Apps", "/marketplace/connectors/typeform", "AI builds and analyzes forms", 2, "Existing connector."),
("Beehiiv MCP", "EN", "Transactional", "MOFU", "Newsletter operators", "Phrase", "Connector Apps", "/marketplace/connectors/beehiiv", "Newsletter workflows for AI", 2, "Existing connector."),
("Kit MCP", "EN", "Transactional", "MOFU", "Creators", "Phrase", "Connector Apps", "/marketplace/connectors/kit", "Your newsletter run by AI", 2, "Existing connector."),
("GitHub CLI AI agent", "EN", "Commercial", "MOFU", "Developers", "Phrase", "CLI Tools", "/marketplace/cli-tools/github-cli", "Keep GitHub CLI workflows near your agent", 2, "Existing CLI tool page."),
("Stripe CLI AI agent", "EN", "Commercial", "MOFU", "Developers", "Phrase", "CLI Tools", "/marketplace/cli-tools/stripe-cli", "Stripe CLI workflows for agents", 2, "Existing CLI tool page."),
("Vercel CLI AI agent", "EN", "Commercial", "MOFU", "Developers", "Phrase", "CLI Tools", "/marketplace/cli-tools/vercel-cli", "Vercel CLI with synced terminal context", 2, "Existing CLI tool page."),
("Supabase CLI AI agent", "EN", "Commercial", "MOFU", "Developers", "Phrase", "CLI Tools", "/marketplace/cli-tools/supabase-cli", "Supabase CLI workflows for agents", 2, "Existing CLI tool page."),
("Wrangler CLI AI agent", "EN", "Commercial", "MOFU", "Developers", "Phrase", "CLI Tools", "/marketplace/cli-tools/wrangler", "Cloudflare CLI workflows for agents", 2, "Existing CLI tool page."),
("marketplace MCP", "ES", "Commercial", "MOFU", "MCP ES", "Phrase", "MCP ES", "/es/marketplace", "MCP servers para Claude, Codex y Gemini", 1, "Core ES."),
("marketplace de MCP servers", "ES", "Commercial", "MOFU", "MCP ES", "Phrase", "MCP ES", "/es/marketplace", "Encontrá e instalá conectores MCP", 1, "Core ES."),
("conectores para agentes IA", "ES", "Commercial", "MOFU", "Dev ES", "Phrase", "Connectors ES", "/es/marketplace/connectors", "Conectá tu agente a tus apps", 1, "Core ES."),
("skills para Claude Code", "ES", "Commercial", "MOFU", "Claude ES", "Phrase", "Skills ES", "/es/marketplace/skills", "Skills reutilizables para Claude", 1, "Core ES."),
("skills para Codex", "ES", "Commercial", "MOFU", "Codex ES", "Phrase", "Skills ES", "/es/marketplace/skills", "Skills reutilizables para Codex", 1, "Core ES."),
("publicar MCP server", "ES", "Transactional", "BOFU", "Creators ES", "Phrase", "Publishers ES", "/es/marketplace/submit", "Publicá tu MCP server", 1, "Supply ES."),
("vender herramientas IA para developers", "ES", "Transactional", "BOFU", "Creators ES", "Phrase", "Publishers ES", "/es/marketplace/submit", "Vendé herramientas a developers con agentes IA", 2, "Supply ES."),
("Google Drive MCP español", "ES", "Transactional", "MOFU", "Knowledge workers ES", "Phrase", "Connector Apps ES", "/es/marketplace/connectors/google-drive", "Tu Drive conectado al agente", 2, "App-specific ES."),
("Notion MCP español", "ES", "Transactional", "MOFU", "Teams ES", "Phrase", "Connector Apps ES", "/es/marketplace/connectors/notion", "Notion dentro de tu agente IA", 2, "App-specific ES."),
("Gmail MCP español", "ES", "Transactional", "MOFU", "Operators ES", "Phrase", "Connector Apps ES", "/es/marketplace/connectors/gmail", "Tu inbox conectado a IA", 2, "App-specific ES."),
]

wb = Workbook()
ws1 = wb.active
ws1.title = "TerminalSync.ai Keywords"
ws2 = wb.create_sheet("Marketplace Keywords")

brand_fill = PatternFill("solid", fgColor="1E40AF")
sub_fill = PatternFill("solid", fgColor="DBEAFE")
header_fill = PatternFill("solid", fgColor="0F172A")
white = "FFFFFF"
thin = Side(style="thin", color="CBD5E1")
priority1 = PatternFill("solid", fgColor="DCFCE7")
priority2 = PatternFill("solid", fgColor="FEF9C3")
priority3 = PatternFill("solid", fgColor="FFEDD5")

for ws, rows, title in [(ws1, product_rows, "Keywords para publicitar TerminalSync.ai"), (ws2, market_rows, "Keywords para TerminalSync Marketplace")]:
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16, color=white)
    ws["A1"].fill = brand_fill
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A2"] = "Prioridad 1 = campañas de búsqueda primero; Prioridad 2 = test/SEO; Prioridad 3 = broad/competidor. Fuente: investigación web ligera + contenido actual del repo + inferencia de intención de búsqueda."
    ws["A2"].font = Font(italic=True, color="334155")
    ws["A2"].fill = sub_fill
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = Font(bold=True, color=white)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for r, row in enumerate(rows, 4):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    last_row = 3 + len(rows)
    table = Table(displayName=("TerminalSyncKeywords" if ws is ws1 else "MarketplaceKeywords"), ref=f"A3:{get_column_letter(len(headers))}{last_row}")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{last_row}"
    widths = [34, 10, 18, 10, 22, 18, 22, 28, 42, 10, 38]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 42
    for rr in range(4, last_row + 1):
        ws.row_dimensions[rr].height = 34
    ws.conditional_formatting.add(f"J4:J{last_row}", CellIsRule(operator="equal", formula=["1"], fill=priority1))
    ws.conditional_formatting.add(f"J4:J{last_row}", CellIsRule(operator="equal", formula=["2"], fill=priority2))
    ws.conditional_formatting.add(f"J4:J{last_row}", CellIsRule(operator="equal", formula=["3"], fill=priority3))
    dv = DataValidation(type="list", formula1='"Exact,Phrase,Broad"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(f"F4:F{last_row}")
    for col in ["A", "H", "I", "K"]:
        for cell in ws[col][3:last_row]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

out = Path("outputs/keyword-research/terminalsync_keyword_research.xlsx")
wb.save(out)
print(out.resolve())
print(len(product_rows), len(market_rows))
