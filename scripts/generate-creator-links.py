#!/usr/bin/env python3
"""
Genera URLs de outreach para cada creator del tracker xlsx.

Lee:    ~/Downloads/terminalsync-creator-tracker.xlsx (sheet "Lista de creators")
Escribe: ./outreach-links.csv
         columnas: #, creator, handle, language, tier, priority, channel, landing, url

Uso:    python3 scripts/generate-creator-links.py

El campo "Plataforma" del xlsx se mapea a un canal canónico para utm_source.
"YouTube + X" → youtube (la surface de mayor confianza para un cold pitch).
Los X-first (Riley Brown, McKay) van como "x". Editable per-row en el CSV.
"""
from __future__ import annotations
import csv
import re
import sys
from pathlib import Path
from urllib.parse import urlencode

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl missing. Install with: pip3 install openpyxl")

XLSX = Path.home() / "Downloads" / "terminalsync-creator-tracker.xlsx"
ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "outreach-links.csv"

CAMPAIGN = "launch-2026-q2"
ORIGIN = "https://terminalsync.ai"


def creator_utm(handle: str, channel: str, lang: str = "es", landing: str = "dev",
                variant: str | None = None, campaign: str = CAMPAIGN) -> str:
    h = handle.lstrip("@").lower().strip()
    path = f"/{lang}/for-developers" if landing == "dev" else f"/{lang}"
    params = {
        "utm_source": channel,
        "utm_medium": "creator",
        "utm_campaign": campaign,
        "utm_term": h,
        "ref": h,
    }
    if variant:
        params["utm_content"] = variant
    return f"{ORIGIN}{path}?{urlencode(params)}"


def pick_channel(plataforma: str) -> str:
    p = (plataforma or "").lower()
    if "x (principal)" in p or p.startswith("x +"):
        return "x"
    if "youtube" in p:
        return "youtube"
    if "newsletter" in p:
        return "newsletter"
    if "podcast" in p:
        return "podcast"
    if "reddit" in p:
        return "reddit"
    if "discord" in p or "slack" in p:
        return "discord"
    if "web" in p:
        return "email"
    return "email"


def derive_handle(creator_name: str, url_canal: str | None) -> str:
    if url_canal:
        m = re.search(r"@([\w.\-]+)", url_canal) or re.search(r"/([\w.\-]+)/?$", url_canal)
        if m:
            return m.group(1)
    return re.sub(r"[^a-z0-9]", "", creator_name.lower())


def main() -> None:
    if not XLSX.exists():
        sys.exit(f"xlsx not found: {XLSX}")
    wb = load_workbook(XLSX, data_only=True)
    ws = wb["Lista de creators"]
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}

    rows_out: list[list[object]] = [
        ["#", "creator", "handle", "language", "tier", "priority", "channel", "landing", "url"]
    ]
    for row in ws.iter_rows(min_row=2, values_only=True):
        creator = row[idx.get("Creator", 1)]
        if not creator:
            continue
        idioma = (row[idx.get("Idioma", 2)] or "").lower()
        lang = "en" if idioma == "en" else "es"
        plataforma = row[idx.get("Plataforma", 4)]
        url_canal = row[idx.get("URL canal", 5)]
        channel = pick_channel(plataforma)
        handle = derive_handle(creator, url_canal)
        landing = "dev"  # default for the entire 50-creator list
        url = creator_utm(handle, channel, lang, landing)
        rows_out.append([
            row[idx.get("#", 0)],
            creator,
            handle,
            lang,
            row[idx.get("Tier", 3)],
            row[idx.get("Prioridad", 12)],
            channel,
            landing,
            url,
        ])

    with OUT.open("w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(rows_out)
    print(f"Wrote {len(rows_out) - 1} rows → {OUT}")


if __name__ == "__main__":
    main()
