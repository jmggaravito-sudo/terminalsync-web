#!/usr/bin/env python3
"""
TSync — manual outreach sender (curated 50).

Reads the hand-curated tracker xlsx + the generated UTM links, picks creators
by handle/range, renders the launch-plan templates (EN/ES) with personalized
{Name} + UTM link, and sends via Resend (terminalsync.ai).

Use this for the FIRST batches you fire by hand. The Resend API is called
directly — no n8n round-trip.

Usage:
  # 1. Dry-run a single creator (prints the rendered email, sends nothing):
  python3 scripts/send-outreach.py --handle ColeMedin --dry-run

  # 2. Real send to a single creator (asks for explicit y/N confirm):
  python3 scripts/send-outreach.py --handle ColeMedin --to cole@ottomator.ai

  # 3. Batch dry-run for the top 5 priority-1 EN:
  python3 scripts/send-outreach.py --priority 1 --lang en --limit 5 --dry-run

Required env (read from ~/projects/terminalsync-web/.env.local OR shell):
  RESEND_API_KEY   — Resend TSync key (re_NH6n…)
  FROM_EMAIL       — defaults to "Juan <jgaravito@terminalsync.ai>"

The xlsx and outreach-links.csv must exist:
  ~/Downloads/terminalsync-creator-tracker.xlsx (or -FIXED.xlsx)
  ~/projects/terminalsync-web/outreach-links.csv  (run generate-creator-links.py first)
"""
from __future__ import annotations
import argparse
import csv
import os
import re
import sys
from pathlib import Path
from typing import Iterable

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl missing. Install: pip3 install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
ENV_FILE = ROOT / ".env.local"
DOWNLOADS = Path.home() / "Downloads"
LINKS_CSV = ROOT / "outreach-links.csv"

# Prefer the FIXED xlsx if present (4 dead handles updated)
TRACKER = (
    DOWNLOADS / "terminalsync-creator-tracker-FIXED.xlsx"
    if (DOWNLOADS / "terminalsync-creator-tracker-FIXED.xlsx").exists()
    else DOWNLOADS / "terminalsync-creator-tracker.xlsx"
)

# ----------------------------------------------------------------------------
# Templates from the launch plan (sections 3.1 EN cold + 3.2 ES cold).
# {NAME}, {VIDEO_REF}, {INSIGHT}, {PAIN}, {LOOM}, {OFFER_AMOUNT}, {URL}
# Pricing default per tier — Micro=$400, Medium=$1500, others=$0 (affiliate).
# ----------------------------------------------------------------------------

SUBJECTS = {
    "en": "Resurrection for your Claude Code sessions (60s demo inside)",
    "es": "Resurrección de sesiones para Claude Code (demo de 60s)",
}

TEMPLATES = {
    "en": """Hey {NAME},

Saw your video on {VIDEO_REF} — sharp take on {INSIGHT}.
The pain you described about {PAIN} is exactly what I've been
building for the last 8 months.

TerminalSync is the only sync layer that:
- Resurrects Claude Code / Codex / Gemini sessions after crash, sleep,
  or internet drop
- Shares memory across all three agents on the same project
- Encrypts everything AES-256 zero-knowledge before it leaves your Mac

60-sec demo (no signup): {URL}
{LOOM_LINE}

If you're up for covering it, here's what I can offer:

1. Lifetime Pro account ($228/year value)
2. ${OFFER_AMOUNT} for an integrated 8-15 min video
3. 30% recurring lifetime via our affiliate program
4. $200 cash bonus when your link hits 10 paid signups

Happy to send early access today, no commitment. If after 2 weeks it's
not what you expected, no hard feelings.

Worth a quick reply?

Juan
TerminalSync.ai
""",
    "es": """Hola {NAME},

Vi tu video sobre {VIDEO_REF} — buena observación sobre {INSIGHT}.
El dolor que describiste de {PAIN} es exactamente lo que vengo
construyendo hace 8 meses.

TerminalSync es la única capa de sync que:
- Resucita sesiones de Claude Code / Codex / Gemini si se cae internet,
  dormís el Mac, o cerrás sin querer
- Comparte memoria entre los tres agentes en el mismo proyecto
- Cifra todo AES-256 zero-knowledge antes de salir de tu Mac

Demo de 60s (sin signup): {URL}
{LOOM_LINE}

Si te llama la atención cubrirlo, te ofrezco:

1. Cuenta Pro lifetime (vale $228/año)
2. ${OFFER_AMOUNT} por un video integrado de 8-15 min
3. 30% recurrente lifetime por el programa de afiliados
4. Bonus $200 cash cuando tu link llegue a 10 ventas pagas

Te puedo dar acceso early hoy mismo, sin compromiso. Si después de 2
semanas no te convence, no pasa nada.

¿Te interesa que charlemos?

Juan
TerminalSync.ai
""",
}

TIER_OFFER = {
    "Micro": 400,
    "Medium": 1500,
    "High": 0,  # affiliate-only — too expensive for paid
    "Niche": 0,
    "Comm": 0,
    "News": 1500,
}


def load_env() -> dict[str, str]:
    env = dict(os.environ)
    if ENV_FILE.exists():
        for line in ENV_FILE.read_text().splitlines():
            if "=" in line and not line.lstrip().startswith("#"):
                k, v = line.split("=", 1)
                env.setdefault(k.strip(), v.strip().strip('"'))
    return env


def load_creators() -> list[dict]:
    if not TRACKER.exists():
        sys.exit(f"Tracker xlsx not found: {TRACKER}")
    wb = load_workbook(TRACKER, data_only=True)
    ws = wb["Lista de creators"]
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        out.append({h: row[i] for i, h in enumerate(headers)})
    return out


def load_links() -> dict[str, str]:
    if not LINKS_CSV.exists():
        sys.exit(f"outreach-links.csv missing — run scripts/generate-creator-links.py first")
    out: dict[str, str] = {}
    with LINKS_CSV.open() as f:
        for row in csv.DictReader(f):
            out[(row.get("handle") or "").lower()] = row.get("url", "")
    return out


def derive_handle(creator_name: str, url_canal: str | None) -> str:
    if url_canal:
        m = re.search(r"@([\w.\-]+)", url_canal) or re.search(r"/([\w.\-]+)/?$", url_canal)
        if m:
            return m.group(1).lower()
    return re.sub(r"[^a-z0-9]", "", creator_name.lower())


def render(creator: dict, links: dict[str, str], loom_url: str | None) -> tuple[str, str, str]:
    """Returns (subject, html, plain_text)."""
    lang = "en" if (creator.get("Idioma", "") or "").lower() == "en" else "es"
    name = (creator.get("Creator", "").split()[0] or "there").strip()
    handle = derive_handle(creator.get("Creator", ""), creator.get("URL canal", ""))
    url = links.get(handle.lower(), f"https://terminalsync.ai/{lang}/for-developers")

    tier = creator.get("Tier", "")
    offer = TIER_OFFER.get(tier, 0)

    subject = SUBJECTS[lang]
    template = TEMPLATES[lang]
    loom_line = f"Personalized 90s pitch: {loom_url}" if loom_url else ""

    text = (
        template
        .replace("{NAME}", name)
        .replace("{VIDEO_REF}", "[VIDEO/POST específico]")
        .replace("{INSIGHT}", "[insight específico]")
        .replace("{PAIN}", "[contexto perdido / sesión que muere / cambiar de Mac]")
        .replace("{LOOM_LINE}", loom_line)
        .replace("{OFFER_AMOUNT}", str(offer))
        .replace("{URL}", url)
    )

    # Simple HTML rendering — paragraphs from text + a clickable URL.
    paragraphs = [p.strip() for p in text.strip().split("\n\n") if p.strip()]
    html = "\n".join(f"<p>{p.replace(chr(10), '<br>')}</p>" for p in paragraphs)
    html = html.replace(url, f'<a href="{url}">{url}</a>')

    return subject, html, text


def send_via_resend(api_key: str, from_email: str, to: str, subject: str, html: str, reply_to: str | None) -> tuple[int, str]:
    import urllib.request
    import json
    payload = {
        "from": from_email,
        "to": [to],
        "subject": subject,
        "html": html,
    }
    if reply_to:
        payload["reply_to"] = reply_to
    req = urllib.request.Request(
        "https://api.resend.com/emails",
        data=json.dumps(payload).encode(),
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            return resp.status, resp.read().decode()
    except urllib.error.HTTPError as e:
        return e.code, e.read().decode()


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--handle", help="single creator handle to send to (case-insensitive)")
    ap.add_argument("--priority", type=int, choices=[1, 2, 3], help="filter by Prioridad column")
    ap.add_argument("--lang", choices=["en", "es"], help="filter by Idioma")
    ap.add_argument("--tier", help="filter by Tier (Micro/Medium/High/Niche/Comm/News)")
    ap.add_argument("--limit", type=int, default=999)
    ap.add_argument("--dry-run", action="store_true", help="render but don't send")
    ap.add_argument("--to", help="override destination email (for testing)")
    ap.add_argument("--from-email", help="from address; defaults to Juan <jgaravito@terminalsync.ai>")
    ap.add_argument("--loom", help="optional personalized Loom URL to insert into the body")
    args = ap.parse_args()

    env = load_env()
    api_key = env.get("RESEND_API_KEY")
    if not api_key and not args.dry_run:
        sys.exit("RESEND_API_KEY missing — set in .env.local or shell env")

    from_email = args.from_email or env.get("FROM_EMAIL") or "Juan <jgaravito@terminalsync.ai>"
    creators = load_creators()
    links = load_links()

    selected: list[dict] = []
    for c in creators:
        if args.handle:
            ch = derive_handle(c.get("Creator", ""), c.get("URL canal", ""))
            if ch.lower() == args.handle.lower():
                selected.append(c)
        else:
            if args.priority is not None and c.get("Prioridad") != args.priority:
                continue
            if args.lang and (c.get("Idioma", "") or "").lower() != args.lang:
                continue
            if args.tier and c.get("Tier") != args.tier:
                continue
            selected.append(c)
    selected = selected[: args.limit]

    if not selected:
        sys.exit("No creators matched the filters.")

    print(f"\n{'='*72}\n{len(selected)} creator(s) selected:\n")
    for c in selected:
        print(f"  · {c.get('#'):>2}. {c.get('Creator'):28} [{c.get('Idioma','')}/{c.get('Tier','')} P{c.get('Prioridad','')}]")
    print()

    for c in selected:
        subject, html, text = render(c, links, args.loom)
        print(f"\n{'─'*72}")
        print(f"#{c.get('#')} {c.get('Creator')}  →  {args.to or '(creator email — fill in)'}")
        print(f"Subject: {subject}\n")
        print(text)
        print(f"\nHTML preview length: {len(html)} chars")

        if args.dry_run:
            continue

        to = args.to
        if not to:
            print(f"  ⏭  no --to passed; skipping send (use --to email@example.com)")
            continue

        confirm = input(f"  Send to {to}? [y/N] ")
        if confirm.lower() != "y":
            print("  ↳ skipped")
            continue
        status, body = send_via_resend(api_key, from_email, to, subject, html, "jgaravito@terminalsync.ai")
        ok = 200 <= status < 300
        print(f"  {'✅' if ok else '❌'} Resend {status}: {body[:200]}")


if __name__ == "__main__":
    main()
